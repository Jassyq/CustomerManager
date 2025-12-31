import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import re
import io
from difflib import SequenceMatcher

# Page configuration
st.set_page_config(
    page_title="Price Manager",
    layout="wide"
)

# Title and description
st.title("Customer Price Manager")
st.markdown("---")

#styling
try:
    with open('style.css') as f:
        st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)
except FileNotFoundError:
    st.warning("CSS file not found - using default styling")


# Initialize session state
if 'original_file_bytes' not in st.session_state:
    st.session_state.original_file_bytes = None
if 'original_filename' not in st.session_state:
    st.session_state.original_filename = None
if 'customers' not in st.session_state:
    st.session_state.customers = []
if 'products' not in st.session_state:
    st.session_state.products = {}
if 'customer_row_map' not in st.session_state:
    st.session_state.customer_row_map = {}
if 'customer_edits' not in st.session_state:
    st.session_state.customer_edits = {}
if 'data_loaded' not in st.session_state:
    st.session_state.data_loaded = False

def similarity(a, b):
    """Calculate string similarity ratio (0.0 to 1.0)"""
    return SequenceMatcher(None, a, b).ratio()

def find_customer_section_start(ws):
    """Find the row where customer data starts (header row with 序号, 姓名, etc)"""
    for row_idx in range(1, ws.max_row + 1):
        cell1 = ws.cell(row_idx, 1).value
        cell2 = ws.cell(row_idx, 2).value
        
        # Look for the header row: 序号, 姓名, 内容, etc
        if cell1 and cell2:
            if str(cell1).strip() == '序号' and str(cell2).strip() == '姓名':
                return row_idx
    return None

def find_product_section_start(ws):
    """Find the row where product data starts (header row with 商品, 单价, etc)"""
    for row_idx in range(1, ws.max_row + 1):
        cell1 = ws.cell(row_idx, 1).value
        cell2 = ws.cell(row_idx, 2).value
        
        # Look for the header row: 商品, 单价, 数量, 金额
        if cell1 and cell2:
            if str(cell1).strip() == '商品' and str(cell2).strip() == '单价':
                return row_idx
    return None

def load_excel_data(file_bytes):
    """Load Excel data from bytes - NO FILE SAVING"""
    try:
        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        
        # Find section headers
        customer_header_row = find_customer_section_start(ws)
        product_header_row = find_product_section_start(ws)
        
        if not customer_header_row:
            st.error("❌ Cannot find customer list header row!")
            st.info("Expected to find a row with '序号' and '姓名' columns")
            return None, None, None
        
        if not product_header_row:
            st.error("❌ Cannot find product list header row!")
            st.info("Expected to find a row with '商品' and '单价' columns")
            return None, None, None
        
        # Customer data starts after header
        customer_data_start_row = customer_header_row + 1
        
        # Product data starts after header
        product_data_start_row = product_header_row + 1
        
        # Extract customers
        customers = []
        customer_row_map = {}
        row_idx = customer_data_start_row
        
        while row_idx < product_header_row:  # Stop before product section
            seq_num = ws.cell(row_idx, 1).value
            name = ws.cell(row_idx, 2).value
            content = ws.cell(row_idx, 3).value
            phone = ws.cell(row_idx, 5).value
            address = ws.cell(row_idx, 6).value
            
            # Stop if we hit an empty sequence number or name
            if seq_num is None or name is None:
                break
            
            # Convert to strings and clean
            name_str = str(name).strip()
            content_str = str(content).strip() if content else ""
            
            customer_data = {
                'seq': seq_num,
                'name': name_str,
                'content': content_str,
                'phone': phone,
                'address': address
            }
            
            customers.append(customer_data)
            customer_row_map[name_str] = customer_data
            row_idx += 1
        
        # Extract products with prices
        products = {}
        row_idx = product_data_start_row
        
        while row_idx <= ws.max_row:
            product_name = ws.cell(row_idx, 1).value
            price = ws.cell(row_idx, 2).value
            
            if product_name is None or price is None:
                # Check if we've hit the end (empty rows)
                if row_idx > product_data_start_row + 100:  # Safety check
                    break
                row_idx += 1
                continue
            
            try:
                price_float = float(price) if price else 0.0
            except (ValueError, TypeError):
                price_float = 0.0
            
            product_name_str = str(product_name).strip()
            
            # Skip empty products
            if product_name_str:
                products[product_name_str] = {
                    'price': price_float
                }
            
            row_idx += 1
        
        return customers, products, customer_row_map
        
    except Exception as e:
        st.error(f"❌ Error loading file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None, None, None

def parse_customer_items(content_text):
    """Parse the customer's content field to extract items and quantities
    
    Looks for pattern: item_name x quantity
    Separators can be comma or Chinese comma（，）
    """
    if not content_text:
        return []
    
    items = []
    
    # Pattern: Match text up to "x[digits]" followed by comma or end
    # This properly handles Chinese commas inside item names
    pattern = r'(.*?)\s*x\s*(\d+)\s*(?:，|,|$)'
    matches = re.finditer(pattern, content_text)
    
    for match in matches:
        item_name = match.group(1).strip()
        qty_str = match.group(2)
        
        # Clean up the item name - remove only leading/trailing separators, not parentheses
        # This preserves closing parentheses like ） at the end of product names
        item_name = re.sub(r'^[\s,，。；]+', '', item_name)
        item_name = re.sub(r'[\s,，。；]+$', '', item_name)
        
        # Skip empty or invalid items
        if not item_name:
            continue
        
        # Skip if this is the total price line
        if '总价' in item_name:  # ✅ CORRECT CHINESE!
            continue
        
        try:
            qty = int(qty_str)
            items.append({'name': item_name, 'qty': qty})
        except ValueError:
            continue
    
    return items

def get_item_price(customer_name, item_name, fuzzy_threshold=0.70):
    """Get the price for an item using EXACT matching first, then FUZZY matching."""
    
    # *** DEBUG: Show what we're looking for ***
    #st.write(f"DEBUG: Looking for '{item_name}'")
    #st.write(f"DEBUG: In products dict? {item_name in st.session_state.products}")
    
    # Check if there's a custom price for this customer and item
    if customer_name in st.session_state.customer_edits:
        if 'custom_prices' in st.session_state.customer_edits[customer_name]:
            if item_name in st.session_state.customer_edits[customer_name]['custom_prices']:
                return st.session_state.customer_edits[customer_name]['custom_prices'][item_name]
    
    # Try EXACT match in product list (fastest)
    if item_name in st.session_state.products:
        price = st.session_state.products[item_name]['price']
        #st.write(f"DEBUG: EXACT MATCH found! Price: {price}")
        return price
    
    st.write(f"DEBUG: No exact match, trying fuzzy...")
    
    # Try FUZZY match to handle truncated text
    best_match = None
    best_score = 0
    
    for product_name in st.session_state.products.keys():
        score = similarity(item_name, product_name)
        if score > best_score:
            best_score = score
            best_match = product_name
    
    # Return fuzzy match if above threshold
    if best_score >= fuzzy_threshold and best_match:
        return st.session_state.products[best_match]['price']
    
    # Try keyword-based matching as last resort
    item_words = set(item_name.split())
    best_keyword_match = None
    best_keyword_count = 0
    
    for product_name in st.session_state.products.keys():
        product_words = set(product_name.split())
        shared_words = item_words & product_words
        
        if len(shared_words) > best_keyword_count:
            best_keyword_count = len(shared_words)
            best_keyword_match = product_name
    
    # If we found a keyword match with at least 2 shared words, use it
    if best_keyword_match and best_keyword_count >= 2:
        return st.session_state.products[best_keyword_match]['price']
    
    # Not found - return 0.0
    return 0.0

def debug_item_lookup(item_name):
    """Debug function to show why a price lookup fails."""
    # Check exact match
    if item_name in st.session_state.products:
        price = st.session_state.products[item_name]['price']
        return {
            'found': True,
            'type': 'exact',
            'price': price,
            'message': f'✅ EXACT MATCH in product list: ${price:.2f}'
        }
    
    # Try fuzzy match
    best_match = None
    best_score = 0
    
    for product_name in st.session_state.products.keys():
        score = similarity(item_name, product_name)
        if score > best_score:
            best_score = score
            best_match = product_name
    
    # Check if fuzzy match is good (70% threshold)
    if best_score >= 0.70:
        price = st.session_state.products[best_match]['price']
        return {
            'found': True,
            'type': 'fuzzy',
            'price': price,
            'score': best_score,
            'matched_name': best_match,
            'message': f'✅ FUZZY MATCH (similarity: {best_score:.1%}): ${price:.2f}'
        }
    
    # Try keyword matching
    item_words = set(item_name.split())
    best_keyword_match = None
    best_keyword_count = 0
    
    for product_name in st.session_state.products.keys():
        product_words = set(product_name.split())
        shared_words = item_words & product_words
        
        if len(shared_words) > best_keyword_count:
            best_keyword_count = len(shared_words)
            best_keyword_match = product_name
    
    # Return keyword match if found
    if best_keyword_match and best_keyword_count >= 2:
        price = st.session_state.products[best_keyword_match]['price']
        return {
            'found': True,
            'type': 'keyword',
            'price': price,
            'keyword_count': best_keyword_count,
            'matched_name': best_keyword_match,
            'message': f'✅ KEYWORD MATCH ({best_keyword_count} shared words): ${price:.2f}'
        }
    
    # Not found - find similar products
    similar_products = []
    
    item_words = set(item_name.split())
    
    for product_name in st.session_state.products.keys():
        product_words = set(product_name.split())
        shared_words = item_words & product_words
        
        if len(shared_words) >= 1:
            similar_products.append({
                'name': product_name,
                'price': st.session_state.products[product_name]['price'],
                'match_words': len(shared_words),
                'total_words': len(product_words)
            })
    
    # Sort by most similar
    similar_products.sort(key=lambda x: x['match_words'], reverse=True)
    
    return {
        'found': False,
        'type': 'not_found',
        'price': 0.0,
        'message': f'❌ NO MATCH in product list',
        'searched_for': item_name,
        'similar': similar_products[:3]
    }

def calculate_total(items, customer_name=None):
    """Calculate total price for items"""
    total = 0.0
    for item in items:
        item_name = item['name']
        qty = item['qty']
        price = get_item_price(customer_name, item_name)
        total += price * qty
    return total

def get_current_items(customer_name):
    """Get current items for a customer (edited or original)"""
    # If we have edits, use those
    if customer_name in st.session_state.customer_edits:
        if 'items' in st.session_state.customer_edits[customer_name]:
            return st.session_state.customer_edits[customer_name]['items']
    
    # Otherwise parse from original content
    if customer_name in st.session_state.customer_row_map:
        content = st.session_state.customer_row_map[customer_name]['content']
        return parse_customer_items(content)
    
    return []

def save_customer_edits(customer_name, items, custom_prices):
    """Save customer edits to session state"""
    if customer_name not in st.session_state.customer_edits:
        st.session_state.customer_edits[customer_name] = {}
    
    st.session_state.customer_edits[customer_name]['items'] = items
    st.session_state.customer_edits[customer_name]['custom_prices'] = custom_prices
    st.session_state.customer_edits[customer_name]['last_modified'] = datetime.now()

def create_export_excel():
    """Create a clean, formatted Excel export with all customer data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Orders"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # Seq
    ws.column_dimensions['B'].width = 15  # Name
    ws.column_dimensions['C'].width = 50  # Items
    ws.column_dimensions['D'].width = 12  # Total
    ws.column_dimensions['E'].width = 15  # Phone
    ws.column_dimensions['F'].width = 30  # Address
    
    # Create header row
    headers = ['序号', '姓名', '商品内容及数量', '总金额', '手机号码', '收货地址']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Fill data rows
    row_idx = 2
    for customer in st.session_state.customers:
        customer_name = customer['name']
        
        # Get current items (edited or original)
        items = get_current_items(customer_name)
        
        # Build items detail string
        items_detail = []
        for item in items:
            price = get_item_price(customer_name, item['name'])
            subtotal = price * item['qty']
            items_detail.append(f"{item['name']} x{item['qty']} (${subtotal:.2f})")
        
        items_text = '\n'.join(items_detail)
        
        # Calculate total
        total = calculate_total(items, customer_name)
        
        # Write row
        ws.cell(row=row_idx, column=1).value = customer['seq']
        ws.cell(row=row_idx, column=2).value = customer_name
        ws.cell(row=row_idx, column=3).value = items_text
        ws.cell(row=row_idx, column=4).value = f"${total:.2f}"
        ws.cell(row=row_idx, column=5).value = customer['phone']
        ws.cell(row=row_idx, column=6).value = customer['address']
        
        # Apply formatting
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            cell.border = border
        
        row_idx += 1
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# Sidebar for file upload
st.sidebar.header("📁 File Upload")

uploaded_file = st.sidebar.file_uploader(
    "Choose Excel file",
    type=['xlsx', 'xls'],
    help="Upload your customer order Excel file"
)

if uploaded_file is not None:
    # Read file into bytes
    file_bytes = uploaded_file.read()
    
    # Only reload if it's a new file
    if st.session_state.original_file_bytes != file_bytes:
        with st.spinner("Loading Excel file..."):
            customers, products, customer_row_map = load_excel_data(file_bytes)
            
            if customers and products:
                st.session_state.original_file_bytes = file_bytes
                st.session_state.original_filename = uploaded_file.name
                st.session_state.customers = customers
                st.session_state.products = products
                st.session_state.customer_row_map = customer_row_map
                st.session_state.data_loaded = True
                st.session_state.customer_edits = {}  # Reset edits on new file

# Main content
if st.session_state.data_loaded:
    
    # Customer selection at top of sidebar (INSIDE data_loaded block)
    st.sidebar.markdown("---")
    st.sidebar.header("👤 Select Customer")
    
    customer_names = [f"{c['seq']}. {c['name']}" for c in st.session_state.customers]
    selected_customer_idx_sidebar = st.sidebar.selectbox(
        "Customer",
        range(len(customer_names)),
        format_func=lambda x: customer_names[x],
        key="customer_selector"
    )
    
    # Export button at the top
    st.sidebar.markdown("---")
    st.sidebar.header("📥 Export")
    
    export_button = st.sidebar.button("📄 Export Clean Excel", use_container_width=True, type="primary")
    
    if export_button:
        with st.spinner("Creating export file..."):
            export_bytes = create_export_excel()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_filename = f"Customer_Orders_Export_{timestamp}.xlsx"
            
            st.sidebar.download_button(
                label="⬇️ Download Export File",
                data=export_bytes,
                file_name=export_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.sidebar.success("✅ Export ready! Click Download button above.")
    
    # Show stats
    total_customers = len(st.session_state.customers)
    edited_customers = len(st.session_state.customer_edits)
    st.sidebar.info(f"👥 Total customers: {total_customers}\n✏️ Edited: {edited_customers}")
    
    # Get selected customer from sidebar selection
    selected_customer = st.session_state.customers[selected_customer_idx_sidebar]
    customer_name = selected_customer['name']
    
    # Check if customer has been edited
    has_edits = customer_name in st.session_state.customer_edits
    
    # Grand Total
    grand_total = 0.0
    for customer in st.session_state.customers:
        cust_name = customer['name']
        items = get_current_items(cust_name)
        customer_total = calculate_total(items, cust_name)
        grand_total += customer_total
        
    #display grand total
    st.markdown(f"<p style='font-size: 24px;'><strong>📊 Grand Total: ${grand_total:,.2f}</strong> </p>", unsafe_allow_html=True)
    st.caption(f"{len(st.session_state.customers)} customers total")
    st.markdown("---")
    
    # Display customer information
    st.subheader("👤 Customer Information")
    st.write(f"**Name:** {customer_name}" f",  **Phone:** {selected_customer['phone']}")
    st.write(f"**Address:** {selected_customer['address']}")
    if has_edits:
        last_modified = st.session_state.customer_edits[customer_name].get('last_modified')
        if last_modified:
            st.info(f"✏️ Last edited: {last_modified.strftime('%Y-%m-%d %H:%M:%S')}")
    
    st.markdown("---")
    
    # Edit items section
    st.subheader("✏️ Edit Order Items & Prices")
    
    # Get current items
    current_items = get_current_items(customer_name)
    
    # Get current custom prices
    current_custom_prices = {}
    if customer_name in st.session_state.customer_edits:
        if 'custom_prices' in st.session_state.customer_edits[customer_name]:
            current_custom_prices = st.session_state.customer_edits[customer_name]['custom_prices'].copy()
    
    st.write("**Current Items:**")
    
    # Column headers
    col_headers = st.columns([3, 1.2, 1, 1.2, 1])
    with col_headers[0]:
        st.write("**Product**")
    with col_headers[1]:
        st.write("**Price ($/unit)**")
    with col_headers[2]:
        st.write("**Qty**")
    with col_headers[3]:
        st.write("**Subtotal**")
    with col_headers[4]:
        st.write("**Del**")
    
    edited_items = []
    custom_price_updates = {}
    
    # Display existing items with DYNAMIC PRICING
    dynamic_total = 0.0
    
    for idx, item in enumerate(current_items):
        # Use item name in key to make it unique per item (prevents inheritance bug)
        item_key = f"{customer_name}_{item['name'].replace(' ', '_')}_{idx}"

        col1, col2, col3, col4, col5 = st.columns([3, 1.2, 1, 1.2, 1])

        # Check delete status first
        delete = st.session_state.get(f"del_{item_key}", False)

        with col1:
            st.text(item['name'])
        
        with col2:
            current_price = get_item_price(customer_name, item['name'])
            
            new_price = st.number_input(
                f"Price###{idx}",
                min_value=0.0,
                value=float(current_price),
                step=0.01,
                format="%.2f",
                key=f"price_{item_key}",
                label_visibility="collapsed"
            )
            
            if abs(new_price - current_price) > 0.001:
                custom_price_updates[item['name']] = new_price
        
        with col3:
            new_qty = st.number_input(
                f"Qty###{idx}",
                min_value=0,
                value=item['qty'],
                step=1,
                key=f"qty_{item_key}",
                label_visibility="collapsed"
            )
        
        with col4:
            if not delete and new_qty > 0:
                price_to_use = new_price
                subtotal = price_to_use * new_qty
                st.text(f"${subtotal:.2f}")
                dynamic_total += subtotal
            else:
                st.text("--")
        
        with col5:
            st.checkbox("Del", key=f"del_{item_key}", label_visibility="collapsed")
        
        # Only add to edited_items if not deleted and quantity > 0
        if not delete and new_qty > 0:
            edited_items.append({'name': item['name'], 'qty': new_qty})
    
    st.markdown("---")

    #adding items and current total
    col_addItem, col_currentTotal = st.columns([2.5, 2])

    with col_addItem:
        searchBar, quantityBar = st.columns([2.8, 0.7])
        
        with searchBar:
            # Search/Select box - dynamically filters as you type
            available_products = list(st.session_state.products.keys())
        
            # Get current search term from session state to maintain filtering
            if 'search_filter' not in st.session_state:
                st.session_state.search_filter = ""
            
            new_item = st.selectbox(
                "Select Product",
                [""] + available_products,
                key=f"new_item_select_{customer_name}",
            )
    
        with quantityBar:
            new_item_qty = st.number_input(
                "Quantity",
                min_value=0,
                value=0,
                step=1,
                key=f"new_item_qty_{customer_name}"
            )
        # Subtotal display (below the selection row)
        if new_item:
            new_item_price = st.session_state.products.get(new_item, {}).get('price', 0.0)
            new_item_subtotal = new_item_price * new_item_qty
            st.markdown(f"Subtotal: ${new_item_subtotal:.2f}")
            st.metric("New Total", f"${dynamic_total + new_item_subtotal:.2f}")
            
            
        # Add to edited items
        if new_item and new_item_qty > 0:
            existing = False
            for item in edited_items:
                if item['name'] == new_item:
                    item['qty'] += new_item_qty
                    existing = True
                    break
                
            if not existing:
                    edited_items.append({'name': new_item, 'qty': new_item_qty})

    with col_currentTotal:
        st.markdown("## 💵 Current Total:")
        st.markdown(f"<p class='current-total-amount'>${dynamic_total:,.2f}</p>", unsafe_allow_html=True)
        st.metric("Total Items", len([item for item in edited_items if item['qty'] > 0]))
        if has_edits:
            st.success("✅ Changes saved in memory")

    # Update custom prices (outside columns)
    current_custom_prices.update(custom_price_updates)

    # Show custom prices if any (outside columns, below everything)
    if custom_price_updates:
        with st.expander("🛠️ Custom Prices for This Customer"):
            for prod, price in custom_price_updates.items():
                base_price = st.session_state.products.get(prod, {}).get('price', 0.0)
                diff = price - base_price
                if diff > 0:
                    st.write(f"**{prod}**: ${base_price:.2f} → ${price:.2f} (+${diff:.2f})")
                elif diff < 0:
                    st.write(f"**{prod}**: ${base_price:.2f} → ${price:.2f} (${diff:.2f})")
        
    st.markdown("---")
    

    # Submit button
    col1, col2 = st.columns([3, 1])  # Left column is wider (empty space)
    with col1:
        pass  # Empty - creates space on the left
        
    with col2:
        if st.button("💾 Save", use_container_width=True, type="primary"):
            # Save edits to session state
            save_customer_edits(customer_name, edited_items, current_custom_prices)
            st.success("✅ Changes saved to memory!")
            st.info("💡 Click 'Export Clean Excel' in sidebar when all edits are done")
            st.rerun()

else:
    # Welcome screen
    st.info("📤 Please upload an Excel file to get started")
    
    st.markdown("""
    ### Upload a file by clicking the button to the left:
    
    **Workflow:**
    1. **Upload** your Excel file 
    2. **Edit** customers one by one
    3. **Save to Memory** after each customer 
    4. **Export** when done - creates clean, formatted Excel

    ### 📊 Required File Format:
    
    Your Excel file must have:
    - A header row with "序号" and "姓名" (customer list marker)
    - Followed by customer data rows
    - A header row with "商品" and "单价" (product list marker)
    - Followed by product data rows with: Product Name | Price
    
    **Your file structure:**
    - Row 4: Header (序号, 姓名, 内容, 标签, 手机号码, 收货地址)
    - Rows 5+: Customer data
    - Row 40: Product header (商品, 单价, 数量, 金额)
    - Rows 41+: Product data (商品名, 价格)
    """)

# Footer
st.caption("version 1.7.0")
